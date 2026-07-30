import json
import os
from concurrent.futures import ThreadPoolExecutor
from time import time
from typing import Dict, List, Any

import cloudscraper
import openpyxl
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, BarColumn,
    TextColumn, TimeRemainingColumn
)
from rich.table import Table

# ============ 配置 ============
EXCEL_PATH = "合月亭目录.xlsx"
LAST_RESULT_FILE = "last_result.json"
MAX_WORKERS = 20
TIMEOUT = 15
# ==============================


class SiteChecker:
    def __init__(self):
        self.console = Console()
        self.sites = []

    def _parse_excel(self) -> List[Dict[str, str]]:
        sites = []
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    name = str(row[0]).strip() if row[0] else ""
                    url = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    if not url:
                        continue
                    if not url.startswith(("http://", "https://")):
                        url = "https://" + url
                    if not name:
                        name = url
                    sites.append({"name": name, "url": url, "section": sheet_name})
            wb.close()
            return sites
        except FileNotFoundError:
            self.console.print(f"[bold red]找不到 '{EXCEL_PATH}'，请确认文件在同目录下[/bold red]")
            input("按回车退出...")
            exit(1)

    def _show_sections(self):
        sections = {}
        for site in self.sites:
            sections.setdefault(site["section"], []).append(site)

        keys = list(sections.keys())
        text = ""
        for i, name in enumerate(keys):
            text += f"[cyan]{i+1}[/cyan]. {name} ({len(sections[name])}个)\n"

        self.console.print(Panel(
            text.strip(),
            title="[bold yellow]选择检测分类[/bold yellow]",
            subtitle=f"共 {len(self.sites)} 个站点",
            border_style="green"
        ))
        return sections, keys

    def _select_sections(self):
        sections, keys = self._show_sections()
        selected = self.console.input("[bold]编号（空格分隔，0=全部，R=刷新列表）:[/bold] > ").strip()

        if selected.upper() == "R":
            return "refresh"

        if selected == "0":
            return self.sites

        result = []
        try:
            for i in selected.split():
                idx = int(i) - 1
                if 0 <= idx < len(keys):
                    result.extend(sections[keys[idx]])
        except ValueError:
            self.console.print("[red]输入无效[/red]")
            return "invalid"

        if not result:
            self.console.print("[yellow]未选择任何分类[/yellow]")
            return "invalid"
        return result

    @staticmethod
    def _check_single_site(site: Dict[str, str]) -> Dict[str, Any]:
        scraper = cloudscraper.create_scraper(browser={'custom': 'Mozilla/5.0'})
        res = {**site, "status": "异常", "result": ""}
        try:
            resp = scraper.get(site["url"], timeout=TIMEOUT, allow_redirects=True)
            if 200 <= resp.status_code < 400:
                res["status"] = "可用"
            res["result"] = str(resp.status_code)
        except Exception as e:
            res["result"] = type(e).__name__
        return res

    def _load_last(self) -> Dict[str, str]:
        if os.path.exists(LAST_RESULT_FILE):
            with open(LAST_RESULT_FILE, encoding="utf-8") as f:
                return json.load(f)
        return {}

    def _save_result(self, results: List[Dict[str, Any]]):
        data = {r["url"]: r["status"] for r in results}
        with open(LAST_RESULT_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)

    def _display(self, results: List[Dict[str, Any]], duration: float):
        table = Table(title="[bold]检测报告[/bold]", header_style="bold magenta")
        table.add_column("状态", justify="center")
        table.add_column("分类", style="yellow")
        table.add_column("名称", style="green")
        table.add_column("URL", style="dim")
        table.add_column("结果", justify="right")

        results.sort(key=lambda x: (x["section"], x["status"] == "异常"))

        for r in results:
            icon = "✅" if r["status"] == "可用" else "❌"
            color = "green" if r["status"] == "可用" else "red"
            table.add_row(
                f"[{color}]{icon} {r['status']}[/]",
                r["section"], r["name"], r["url"],
                f"[{color}]{r['result']}[/]"
            )

        self.console.print(table)
        alive = sum(1 for r in results if r["status"] == "可用")
        self.console.print(
            f"\n共 {len(results)} 个 | [green]{alive} 可用[/] | "
            f"[red]{len(results)-alive} 异常[/] | 耗时 {duration:.2f}s"
        )

        last = self._load_last()
        revived = [
            r for r in results
            if r["status"] == "可用" and last.get(r["url"]) == "异常"
        ]
        if revived:
            self.console.print(f"\n[bold green]🎉 本次新复活 {len(revived)} 个站点！[/bold green]")
            for r in revived:
                self.console.print(f"   ✅ {r['name']} → {r['url']}")

        self._save_result(results)

    def _run_check(self, sites: List[Dict[str, str]]):
        self.console.print(f"\n[bold]检测 {len(sites)} 个站点...[/bold]")
        start = time()
        results = []

        with Progress(
            SpinnerColumn(),
            "[progress.description]{task.description}",
            BarColumn(),
            "[progress.percentage]{task.percentage:>3.0f}%",
            TextColumn("({task.completed}/{task.total})"),
            TimeRemainingColumn(),
            console=self.console,
        ) as progress:
            task = progress.add_task("[cyan]检测中...", total=len(sites))
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                futures = [pool.submit(self._check_single_site, s) for s in sites]
                for f in futures:
                    results.append(f.result())
                    progress.update(task, advance=1)

        duration = time() - start
        self._display(results, duration)

    def run(self):
        while True:
            self.sites = self._parse_excel()
            if not self.sites:
                self.console.print("[yellow]表格为空，退出。[/yellow]")
                input("按回车退出...")
                return

            choice = self._select_sections()

            if choice == "refresh":
                self.console.print("[cyan]已刷新列表。[/cyan]\n")
                continue

            if choice == "invalid":
                continue

            self._run_check(choice)

            self.console.input("\n[bold]输入 C 回到列表:[/bold] > ")


if __name__ == "__main__":
    checker = SiteChecker()
    checker.run()
